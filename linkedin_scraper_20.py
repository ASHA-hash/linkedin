import json

from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC, wait
import time
import csv
import openpyxl
import datetime
import re

# File name where the number will be stored
file_name = "counter.txt"

# Try to open the file and read the current number
try:
    with open(file_name, "r") as file:
        number = int(file.read())
except FileNotFoundError:
    # If the file does not exist, start with 0
    number = 0

# Increment the number
number += 1

# Write the new number back to the file
with open(file_name, "w") as file:
    file.write(str(number))

# Print the updated number
print(f"Current number: {number}")


def contains_number(s):
    return any(char.isdigit() for char in s)

def scrape_company_data(url):
    driver.get(url)

    Exp_Org_Headline = ""
    Exp_Org_Industry = ""
    Exp_Org_Location = ""
    Exp_Org_Followers = ""
    Exp_Org_EmpSize = ""
    Exp_Org_About = ""
    Exp_Org_Founded = ""

    # XPath for Organization Headline
    try:
        Exp_Org_Headline_css_selector = 'h1.org-top-card-summary__title'
        Exp_Org_Headline = driver.find_element(By.CSS_SELECTOR, Exp_Org_Headline_css_selector).text
    except NoSuchElementException:
        Exp_Org_Headline = ""
        print("Organization Headline not found")

    try:
        Exp_Org_Industry = driver.find_element(By.CSS_SELECTOR,
                                               'div.org-top-card-summary-info-list__info-item').text
    except NoSuchElementException:
        Exp_Org_Industry = ""
        print("Organization Industry not found")

    try:
        # Use the CSS selector to find the elements
        element_1 = driver.find_elements(By.CSS_SELECTOR,
                                         'div.inline-block > .org-top-card-summary-info-list__info-item')

        if len(element_1) >= 3:
            Exp_Org_Location = element_1[0].text
            Exp_Org_Followers = element_1[1].text.split()[0]
            Exp_Org_EmpSize = element_1[2].text.split()[0]
        else:
            print("Not enough elements found for Location, Followers, or Employee Size")
    except NoSuchElementException:
        Exp_Org_Location = ""
        Exp_Org_Followers = ""
        Exp_Org_EmpSize = ""
        print("Organization details not found")

    try:
        see_more = driver.find_element(By.CSS_SELECTOR, 'a.lt-line-clamp__more')
        see_more.click()
    except:
        pass

    try:
        Exp_Org_About = driver.find_element(By.CSS_SELECTOR, 'div.org-about-module__description').text
    except:
        Exp_Org_About = ""
        print("Organization About section not found")

    url = url + "about/"
    try:
        driver.get(url)
        print("Entered about")
        time.sleep(10)

        # List of possible XPaths
        xpaths = [
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[3]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[4]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[5]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[6]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[7]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[8]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[9]',
            '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[10]'
        ]

        Exp_Org_Founded = None

        # Iterate over each XPath
        for xpath in xpaths:
            try:
                # Try to find the element using the current XPath
                Exp_Org_Founded = driver.find_element(By.XPATH, xpath).text
                print(f"Found with XPath {xpath}: {Exp_Org_Founded}")
                if Exp_Org_Founded:
                    try:
                        # Attempt to convert the text to an integer
                        int_value = int(Exp_Org_Founded)
                        Exp_Org_Founded = int_value
                        print(f"Exp_Org_Founded as integer: {Exp_Org_Founded}")
                        break  # Exit the loop if found
                    except ValueError:
                        # If conversion fails, store it as an empty string or custom message
                        Exp_Org_Founded = "No Exp_Org_Founded"
                        print("Conversion to integer failed, stored as 'No Exp_Org_Founded'")
                else:
                    Exp_Org_Founded = "No Exp_Org_Founded"
                    print("Exp_Org_Founded not found, stored as 'No Exp_Org_Founded'")


            except:
                continue  # If not found, move to the next XPath



    except Exception as e:
        Exp_Org_Founded = ""
        print(f"An error occurred")
    # Printing the extracted values (for verification)
    print(f"Organization Headline: {Exp_Org_Headline}")
    print(f"Industry: {Exp_Org_Industry}")
    print(f"Location: {Exp_Org_Location}")
    print(f"Employee Size: {Exp_Org_EmpSize}")
    print(f"Followers: {Exp_Org_Followers}")
    print(f"About Section: {Exp_Org_About}")
    print(f"Founded Date: {Exp_Org_Founded}")



    # Store extracted values in a dictionary
    company_data = {
        "Exp_Org_Headline": Exp_Org_Headline,
        "Exp_Org_Industry": Exp_Org_Industry,
        "Exp_Org_Location": Exp_Org_Location,
        "Exp_Org_EmpSize": Exp_Org_EmpSize,
        "Exp_Org_Followers": Exp_Org_Followers,
        "Exp_Org_About": Exp_Org_About,
        "Exp_Org_Founded": Exp_Org_Founded,
    }

    # Return the extracted values
    return company_data

# from get_csv import profile_id

file_name = "inputs_links.csv"

# # Read data from the CSV file and extract the links
links = []
with open(file_name, mode='r') as file:
    reader = csv.DictReader(file)

    # Extract only the 'link' field from each row
    for row in reader:
        links.append(row["link"])

# # Output the extracted links
# print(links[6])
# # profile_url = links[6]
# # print(profile_url)
driver = webdriver.Chrome()

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet2 = workbook.create_sheet(title="Activity")
sheet3 = workbook.create_sheet(title="Experience")
sheet4 = workbook.create_sheet(title="Education")
sheet5 = workbook.create_sheet(title="Licenses_&_certifications")
sheet6 = workbook.create_sheet(title="Projects")
sheet7 = workbook.create_sheet(title="Volunteering")
sheet8 = workbook.create_sheet(title="Honors")
sheet9 = workbook.create_sheet(title="Skills")
sheet10 = workbook.create_sheet(title="Recommendations_Received")
sheet11 = workbook.create_sheet(title="Recommendations_Given")
sheet12 = workbook.create_sheet(title="Featured Post")


def linkedIn_login(driver):
    print("Navigating to LinkedIn login page...")
    driver.get("https://www.linkedin.com/")
    time.sleep(5)

    try:

        print("Clicking on the Sign In button...")
        signUpButton_Path = "/html/body/nav/div/a[2]"
        signUpButton = driver.find_element(By.XPATH, signUpButton_Path)
        signUpButton.click()

        time.sleep(3)

        print("Entering email...")
        email_field_Path = "/html/body/div/main/div[2]/div[1]/form/div[1]/input"
        email_field = driver.find_element(By.XPATH, email_field_Path)
        # email_field.click()
        email_field.send_keys("Your_email")
        time.sleep(2)

        print("Entering password...")
        password_field_Path = "/html/body/div/main/div[2]/div[1]/form/div[2]/input"
        password_field = driver.find_element(By.XPATH, password_field_Path)
        # password_field.click()
        password_field.send_keys("Your_Password")
        time.sleep(2)

        print("Clicking on the Sign In button...")
        signInButton_Path = "/html/body/div/main/div[2]/div[1]/form/div[3]/button"
        signInButton = driver.find_element(By.XPATH, signInButton_Path)
        signInButton.click()

        print("Logged in successfully.")

        time.sleep(5)
    except Exception as e:
        print(f"Error during LinkedIn login: {e}")


def profile_basic_data(driver, profile_url, sheet1, row):
    print(f"Navigating to profile URL: {profile_url}")
    driver.get(profile_url)
    sheet1.title = "Profile Basic Data"
    sheet1['A1'] = "Date Of Extraction"
    sheet1['B1'] = "linkedin_url"
    sheet1['C1'] = "profile_id"
    sheet1['D1'] = "profile_id_yn"
    sheet1['E1'] = "fullname"
    sheet1['F1'] = "headline"
    sheet1['G1'] = "location"
    sheet1['H1'] = "num_connections"
    sheet1['I1'] = "highlights"
    sheet1['J1'] = "about"
    try:
        date_of_extraction = datetime.datetime.now().date()
        linkedIn_url = profile_url

        #
        # profile_id = linkedIn_url.rstrip('/').split('/')[-1]
        # profile_id_yn = ""
        # if contains_number(profile_id):
        #     profile_id_yn = 'Y'
        # else:
        #     profile_id_yn = 'N'

        try:
            # Locate the img element using CSS selector
            img_element = driver.find_element(By.CSS_SELECTOR, "button.pv-top-card-profile-picture__container img")

            # Get the outer HTML of the img element
            img_html = img_element.get_attribute("outerHTML")
            profile_id = img_html
            print(profile_id)
            if contains_number(profile_id):
                profile_id_yn = 'Y'
            else:
                profile_id_yn = 'N'
        except:
            profile_id = ""

        print("Extracting profile data...")
        try:
            fullname = driver.find_element(By.CSS_SELECTOR,
                                           'h1.text-heading-xlarge.inline.t-24.v-align-middle.break-words').text
        except:
            fullname = ""
        print(f"fullname : {fullname}")

        try:
            headline = driver.find_element(By.CSS_SELECTOR,
                                           'div.text-body-medium.break-words').text
        except:
            headline = ""
        print(f"headline : {headline}")
        try:
            location = driver.find_element(By.CSS_SELECTOR,
                                           'span.text-body-small.inline.t-black--light.break-words').text
        except:
            location = ""
        print(f"location : {location}")
        try:
            num_connections = driver.find_element(By.CSS_SELECTOR,
                                                  'ul li.text-body-small span.t-black--light span.t-bold').text
        except:
            num_connections = ""
        print(f"num of connections : {num_connections}")
        try:
            highlights_list = []
            # highlights = driver.find_element(By.XPATH,
            #                                  '//*[@id="profile-content"]/div/div[2]/div/div/main/section[2]/div[3]/ul/li/div/div[2]/div[1]/div/div/div/div/div/span[1]').text
            highlights = driver.find_element(By.ID, 'highlights')
            highlights_anchor = highlights.find_element(By.XPATH, "./ancestor::section")
            # Find anchor tags within this section with the specified class
            anchors = highlights_anchor.find_elements(By.XPATH, ".//li[contains(@class, 'artdeco-list__item')]")

            for anchor in anchors:
                # Extract the text or other information you need from each anchor
                text = anchor.text.strip().split("\n")
                text_list = text[1]

                highlights_list.append(text_list)

            print(highlights_list)

            highlights = highlights_list[0]
            print(highlights)

        except NoSuchElementException:
            highlights = ""

        # '//*[@id="profile-content"]/div/div[2]/div/div/main/section[2]/div[3]/div/div/div/span[3]/button'
        # Xpath of the 'See More' button, if it exists
        see_more_button_css_selector = (
            'button.inline-show-more-text__button.inline-show-more-text__button--light.link')

        try:
            # Check if the "See More" button exists and click it if it does
            see_more_button = driver.find_element(By.XPATH, see_more_button_css_selector)
            see_more_button.click()
        except:
            # If the "See More" button doesn't exist, continue
            pass
        finally:
            pass

        try:
            # about_path = driver.find_element(By.CSS_SELECTOR, 'section.pv-profile-card').text
            # about = about_path.find_element(By.CSS_SELECTOR, 'div.inline-show-more-text--is-collapsed').text
            about = driver.find_element(By.ID, 'about')
            about_anchor = about.find_element(By.XPATH, "./ancestor::section")
            # Select the desired div with the specified classes
            about = about_anchor.find_element(By.XPATH, ".//div[@class='display-flex ph5 pv3']").text

            print(f"about : {about}")
        except:
            about = ""

        print("Profile data extracted successfully.")

        sheet1.append(
            [date_of_extraction, linkedIn_url, profile_id, profile_id_yn, fullname, headline, location, num_connections,
             highlights, about])
        time.sleep(5)
        workbook.save("profile_basic_data_7.xlsx")
    except Exception as e:
        print(f"Error during profile data extraction: {e}")

    finally:
        pass


def activity(driver, profile_url, sheet2, row):
    driver.get(profile_url)
    print("Navigating to profile activity page")

    # Set the initial headers in the first row
    headers = ["LinkenIn_URL",
               "Activity_Followers_Count",
               "Activity_Post_Count",
               "Activity_content"
               ]

    sheet2.cell(row=row, column=1, value=profile_url)
    # Initialize the headers list but don't add post-specific headers yet
    try:
        print("Extracting activity data...")
        # Initialize an empty dictionary
        data_list = []

        # Alternative general XPath to locate the follower count based on the structure
        follower_element = driver.find_element(By.XPATH,
                                               "//div[contains(@class, 'pvs-header__top-container')]//p[span[contains(text(), 'followers')]]")

        # Extract the text (follower count) from the elements
        follower_counts = follower_element.text

        print(follower_counts)

        # Extract the number of followers
        Activity_Followers_Count = driver.find_element(By.XPATH,
                                                       "//div[contains(@class, 'pvs-header__top-container')]//p[contains(@class, 'pvs-header__optional-link')]//span[contains(text(), 'followers')]").text
        Activity_Followers_Count = ''.join(filter(str.isdigit, Activity_Followers_Count))
        print(f"Activity followers count: {Activity_Followers_Count}")

        # Generalized XPath to locate the 'a' tag within the footer element
        # show_all_posts_xpath = driver.find_element(By.XPATH,
        #                                     "//footer//a[contains(@class, 'profile-creator-shared-content-view__footer-action')]")
        # show_all_post_link = show_all_posts_xpath.get_attribute("href")
        #
        # print(show_all_post_link)
        show_all_post_link = profile_url + "recent-activity/all/"
        print(show_all_post_link)
        driver.get(show_all_post_link)
        time.sleep(10)

        # Extract the list of posts
        activity_post_list_css_selector = "li.profile-creator-shared-feed-update__container"
        activity_post_list = driver.find_elements(By.CSS_SELECTOR, activity_post_list_css_selector)

        Activity_Post_Count = len(activity_post_list)
        print(f"Number of posts found: {Activity_Post_Count}")

        entries = 6

        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet2.cell(row=1, column=col, value=header)

        # Write Activity_Followers_Count and Activity_Post_Count to the second row

        sheet2.cell(row=row, column=2, value=Activity_Followers_Count)
        sheet2.cell(row=row, column=3, value=Activity_Post_Count)

        # Write data for each post in the second row
        current_col = 4  # Start from column C for post data

        for i, post in enumerate(activity_post_list[:5]):
            try:
                # Locate the container element
                container = driver.find_element(By.CLASS_NAME, 'social-details-social-counts')

                # Find all <li> elements within the container
                list_items = container.find_elements(By.CLASS_NAME, 'social-details-social-counts__item--right-aligned')
                # Check the number of <li> elements and determine the post type
                if len(list_items) == 2:
                    Activity_Post_Type = "reposted"
                else:
                    Activity_Post_Type = "posted"

                print(f"Activity_Post_Type : {Activity_Post_Type}")

            except:
                Activity_Post_Type = "posted"  # Assuming "posted"; update this logic as needed

            try:
                Activity_Post_Date_Num = post.find_element(By.CSS_SELECTOR,
                                                           'span.update-components-actor__sub-description.t-12.t-normal.t-black--light').text
                Activity_Post_Date_Unit = post.find_element(By.CSS_SELECTOR,
                                                            'span.update-components-actor__sub-description.t-12.t-normal.t-black--light').text

                # Find the first number in the string
                # match_ = re.search(r'\d+', Activity_Post_Date_Num)

                # # Extract the number if found
                # if match_:
                #     first_number = int(match_.group())
                #     Activity_Post_Date_Num = first_number
                #     print(first_number)  # Output: 4

                # Use regular expression to extract the number and the unit separately
                match = re.match(r'(\d+)(\D+)', Activity_Post_Date_Num)

                if match:
                    # Extract the number
                    Activity_Post_Date_Num = match.group(1)  # This will give you "4"
                    # Extract the unit
                    Activity_Post_Date_Unit = match.group(2).replace('•', '').strip()  # This will give you "mo"
                else:
                    print("No number found")

                print(Activity_Post_Date_Num)
                print(Activity_Post_Date_Unit)

            except:
                Activity_Post_Date_Num = ""
                Activity_Post_Date_Unit = ""

            # Click "See More" if it exists
            see_more_button_css_selector = 'button.feed-shared-inline-show-more-text__see-more-less-toggle'
            try:
                see_more_button = post.find_element(By.CSS_SELECTOR, see_more_button_css_selector)
                see_more_button.click()
            except:
                pass

            try:
                Activity_Post_Content = post.find_element(By.CSS_SELECTOR,
                                                          'div.feed-shared-update-v2__description-wrapper').text
            except:
                Activity_Post_Content = ""

            try:
                Activity_Post_Reaction_Num = post.find_element(By.CSS_SELECTOR,
                                                               'li.social-details-social-counts__item.social-details-social-counts__reactions').text
            except:
                Activity_Post_Reaction_Num = ""

            try:
                Activity_Post_Comment_Num = post.find_element(By.CSS_SELECTOR,
                                                              'li.social-details-social-counts__comments.social-details-social-counts__item').text.split()[
                    0]
            except:
                Activity_Post_Comment_Num = ""

            print(f"Post {i + 1} data extracted successfully.")

            # Create a dictionary for the current person
            activity_data = {"Activity_Post_Count": i + 1,
                             "Activity_Post_Type": Activity_Post_Type,
                             "Activity_Post_Date_Num": Activity_Post_Date_Num,
                             "Activity_Post_Date_Unit": Activity_Post_Date_Unit,
                             "Activity_Post_Content": Activity_Post_Content,
                             "Activity_Post_Reaction_Num": Activity_Post_Reaction_Num,
                             "Activity_Post_Comment_Num": Activity_Post_Comment_Num}

            # Add the dictionary to the list
            data_list.append(activity_data)
            print(f"{i} data appended")

        print(data_list)

        # Convert the data list to a JSON string
        json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

        # Write the JSON string to the Excel sheet
        sheet2.cell(row=row, column=current_col, value=json_data)
        print("JSON data written to sheet")

        # Save the workbook
        workbook.save("profile_basic_data_7.xlsx")
        print("File saved successfully")



    except Exception as e:

        print("There is no activity found")


def experience(driver, profile_url, sheet3, row):
    print(profile_url)
    driver.get(profile_url)
    print("Experience data are getting extracted...")


    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL","Exp_Count","Experience_Content","Company_details"]

        try:
        # Locate "Show all licenses & certifications" link and navigate to it
            show_all_link = driver.find_element(By.CSS_SELECTOR,
                                                'a#navigation-index-see-all-experiences')
            show_all_url = show_all_link.get_attribute('href')
        except:
            show_all_url = profile_url + "details/experience/"
            print(show_all_url)

        driver.get(show_all_url)

        # Find all license & certification items
        # experience_list = driver.find_elements(By.CSS_SELECTOR, '.pvs-list__paged-list-item, .artdeco-list__item, .pvs-list__item--line-separated, .pvs-list__item--one-column')
        experience_list = driver.find_elements(By.CSS_SELECTOR,
                                               '.pvs-list__paged-list-item.artdeco-list__item.pvs-list__item--line-separated.pvs-list__item--one-column')

        Exp_Count = len(experience_list)
        print(f"Number of experience found: {Exp_Count}")
        entries = 6

        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet3.cell(row=1, column=col, value=header)

        sheet3.cell(row=row, column=1, value=profile_url)
        sheet3.cell(row=row, column=2, value=Exp_Count)
        data_list = []
        # If certificates are found, extract the data
        if Exp_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for Exp_Count)

            a_tags_list = []

            for i, exp in enumerate(experience_list):
                print(i)
                # link = exp.find_element(By.TAG_NAME, 'a')
                # Org_link = link.get_attribute("href")
                Exp_Org_Name = exp.find_element(By.CSS_SELECTOR, '.display-flex .mr1.t-bold span').text.strip()
                print(Exp_Org_Name)

                try:

                    inner_list = exp.find_elements(By.CSS_SELECTOR,
                                                   '.pvs-list__paged-list-item.pvs-list__item--one-column')
                    # print(inner_list.text)
                    print("hellrsedfsdo")
                    inner = exp.find_element(By.CSS_SELECTOR,
                                             '.pvs-list__paged-list-item.pvs-list__item--one-column')

                    Org_Link = exp.find_element(By.TAG_NAME, 'a').get_attribute("href")

                    a_tags_list.append(Org_Link)

                    for a, inn in enumerate(inner_list):
                        entry_text = inn.text
                        # Split the text into lines and store in a list
                        entry_list_1 = entry_text.strip().split("\n")
                        entry_list = []
                        [entry_list.append(item) for item in entry_list_1 if item not in entry_list]
                        print(entry_list)

                        # Extracting and storing information with general error handling

                        try:
                            Exp_Title = entry_list[0].strip()  # Job title
                        except:
                            Exp_Title = None


                        # try:
                        #     Exp_RoleType = "Full-time"  # Assuming role type based on context, adjust if needed
                        # except:
                        #     Exp_RoleType = None

                        try:
                            # Define terms to look for
                            terms_to_find = ['Full-time', 'Contract', 'Self-employed']

                            # Search for terms in the list
                            found_terms = [term for term in terms_to_find if
                                           any(term in content for content in entry_list)]
                            # Store the result in a variable and print it if found
                            if found_terms:
                                Exp_RoleType = found_terms[0]
                                print(Exp_RoleType)

                        except:
                            Exp_RoleType = None

                        try:
                            # Exp_Start = entry_list[3].split(" - ")[0].strip()  # Start date
                            Exp_Start = exp.find_element(By.CSS_SELECTOR,
                                                         'span.t-14.t-normal.t-black--light .pvs-entity__caption-wrapper').text
                            Exp_Start = Exp_Start.split(' - ')[0].strip()
                        except:
                            Exp_Start = None

                        try:
                            # Exp_End = entry_list[3].split(" - ")[1].split("·")[0].strip()  # End date
                            Exp_End = exp.find_element(By.CSS_SELECTOR,
                                                       'span.t-14.t-normal.t-black--light .pvs-entity__caption-wrapper').text
                            Exp_End = Exp_End.split(' - ')[1].strip()
                            Exp_End = Exp_End.split('.')[0]
                        except:
                            Exp_End = None

                        try:
                            # Exp_Location = entry_list[4].strip()  # Full location string
                            Exp_Location = exp.find_elements(By.CSS_SELECTOR,
                                                             'span.t-14.t-normal.t-black--light')
                            Exp_Location_list = []
                            a = 0
                            for i in Exp_Location:
                                print(a)
                                text = i.text
                                # Split the text into lines and strip each line
                                text_list = [line.strip() for line in text.splitlines()]
                                a = a + 1
                                Exp_Location_list.append(text_list[1])
                            print(Exp_Location_list)

                            Exp_Location = Exp_Location_list[1]
                            # Split the string on the delimiter '·'
                            location_parts = Exp_Location.split('·')

                            # Strip any extra whitespace from each part
                            location = location_parts[0].strip()
                            loc_type = location_parts[1].strip() if len(location_parts) > 1 else None
                            Exp_Location = location
                            Exp_LocType = loc_type

                            print(f"Location : {Exp_Location}")
                            print(f"Location Type : {Exp_LocType}")

                        except:
                            Exp_Location = None

                        try:
                            Exp_Desc = max(entry_list, key=len)  # Description
                        except:
                            Exp_Desc = None

                        try:
                            Exp_Org_Link = exp.find_element(By.TAG_NAME, 'a').get_attribute("href")
                            Exp_Org_HasLink = "1" if Exp_Org_Link else "0"

                        except:
                            Exp_Org_Link = ""
                            Exp_Org_HasLink = 0



                        # Print the extracted values (for verification)
                        print(f"Exp_Org_Name: {Exp_Org_Name}")
                        print(f"Exp_Title: {Exp_Title}")
                        print(f"Exp_RoleType: {Exp_RoleType}")
                        print(f"Exp_Start: {Exp_Start}")
                        print(f"Exp_End: {Exp_End}")
                        print(f"Exp_Location: {Exp_Location}")
                        print(f"Exp_LocType: {Exp_LocType}")
                        print(f"Exp_Desc: {Exp_Desc}")
                        print(f"Exp_Org_HasLink: {Exp_Org_HasLink}")
                        print(f"Exp_Org_Link: {Exp_Org_Link}")
                        experience_data = {
                            "Exp_Org_Name": Exp_Org_Name,
                            "Exp_Title": Exp_Title,
                            "Exp_RoleType": Exp_RoleType,
                            "Exp_Start": Exp_Start,
                            "Exp_End": Exp_End,
                            "Exp_Location": Exp_Location,
                            "Exp_LocType": Exp_LocType,
                            "Exp_Desc": Exp_Desc,
                            "Exp_Org_HasLink": Exp_Org_HasLink,
                            "Exp_Org_Link": Exp_Org_Link

                        }
                        # Add the dictionary to the list
                        data_list.append(experience_data)
                        break


                except:

                    print("Billionaire")

                    entry_text = exp.text
                    # Split the text into lines and store in a list
                    entry_list_1 = entry_text.strip().split("\n")
                    entry_list = []
                    [entry_list.append(item) for item in entry_list_1 if item not in entry_list]
                    print(entry_list)

                    Org_Link = exp.find_element(By.TAG_NAME, 'a').get_attribute("href")

                    a_tags_list.append(Org_Link)

                    # Extracting and storing information with general error handling

                    try:
                        Exp_Title = entry_list[0].strip()  # Job title
                    except:
                        Exp_Title = None


                    # try:
                    #     Exp_RoleType = "Full-time"  # Assuming role type based on context, adjust if needed
                    # except:
                    #     Exp_RoleType = None

                    try:
                        # Define terms to look for
                        terms_to_find = ['Full-time', 'Contract', 'Self-employed']

                        # Search for terms in the list
                        found_terms = [term for term in terms_to_find if
                                       any(term in content for content in entry_list)]
                        # Store the result in a variable and print it if found
                        if found_terms:
                            Exp_RoleType = found_terms[0]
                            print(Exp_RoleType)
                        else:
                            Exp_RoleType = None

                    except:
                        Exp_RoleType = None

                    try:
                        # Exp_Start = entry_list[3].split(" - ")[0].strip()  # Start date
                        Exp_Start = exp.find_element(By.CSS_SELECTOR,
                                                     'span.t-14.t-normal.t-black--light .pvs-entity__caption-wrapper').text
                        Exp_Start = Exp_Start.split(' - ')[0].strip()
                    except:
                        Exp_Start = None

                    try:
                        # Exp_End = entry_list[3].split(" - ")[1].split("·")[0].strip()  # End date
                        Exp_End = exp.find_element(By.CSS_SELECTOR,
                                                   'span.t-14.t-normal.t-black--light .pvs-entity__caption-wrapper').text
                        Exp_End = Exp_End.split(' - ')[1].strip()
                        Exp_End = Exp_End.split('.')[0]
                    except:
                        Exp_End = None

                    try:
                        # Exp_Location = entry_list[4].strip()  # Full location string
                        Exp_Location = exp.find_elements(By.CSS_SELECTOR,
                                                         'span.t-14.t-normal.t-black--light')
                        Exp_Location_list = []
                        a = 0
                        for i in Exp_Location:
                            print(a)
                            text = i.text
                            # Split the text into lines and strip each line
                            text_list = [line.strip() for line in text.splitlines()]
                            a = a + 1
                            Exp_Location_list.append(text_list[1])
                        print(Exp_Location_list)

                        Exp_Location = Exp_Location_list[1]
                        # Split the string on the delimiter '·'
                        location_parts = Exp_Location.split('·')

                        # Strip any extra whitespace from each part
                        location = location_parts[0].strip()
                        loc_type = location_parts[1].strip() if len(location_parts) > 1 else None
                        Exp_Location = location
                        Exp_LocType = loc_type

                        print(f"Location : {Exp_Location}")
                        print(f"Location Type : {Exp_LocType}")

                    except:
                        Exp_Location = None
                        Exp_LocType = ""

                    try:
                        Exp_Desc = max(entry_list, key=len)  # Description
                    except:
                        Exp_Desc = None

                    try:
                        Exp_Org_Link = exp.find_element(By.TAG_NAME, 'a').get_attribute("href")
                        # Check if the link contains specific keywords and set the Exp_Org_HasLink variable accordingly
                        if Exp_Org_Link and any(
                                keyword in Exp_Org_Link for keyword in ["company", "school", "institution"]):
                            Exp_Org_HasLink = "1"
                        else:
                            Exp_Org_HasLink = "0"


                    except Exception as e:
                        Exp_Org_Link = ""
                        Exp_Org_HasLink = "0"


                    # Print the extracted values (for verification)
                    print(f"Exp_Org_Name: {Exp_Org_Name}")
                    print(f"Exp_Title: {Exp_Title}")
                    print(f"Exp_RoleType: {Exp_RoleType}")
                    print(f"Exp_Start: {Exp_Start}")
                    print(f"Exp_End: {Exp_End}")
                    print(f"Exp_Location: {Exp_Location}")
                    print(f"Exp_LocType: {Exp_LocType}")
                    print(f"Exp_Desc: {Exp_Desc}")
                    print(f"Exp_Org_HasLink: {Exp_Org_HasLink}")
                    print(f"Exp_Org_Link: {Exp_Org_Link}")

                    experience_data = {
                        "Exp_Org_Name": Exp_Org_Name,
                        "Exp_Title": Exp_Title,
                        "Exp_RoleType": Exp_RoleType,
                        "Exp_Start": Exp_Start,
                        "Exp_End": Exp_End,
                        "Exp_Location": Exp_Location,
                        "Exp_LocType": Exp_LocType,
                        "Exp_Desc": Exp_Desc,
                        "Exp_Org_HasLink": Exp_Org_HasLink,
                        "Exp_Org_Link": Exp_Org_Link

                    }
                    # Add the dictionary to the list
                    data_list.append(experience_data)





            company_details = []

            for i, exp in enumerate(experience_list):

                driver.get(a_tags_list[i])

                Exp_Org_Headline = ""
                Exp_Org_Industry = ""
                Exp_Org_Location = ""
                Exp_Org_Followers = ""
                Exp_Org_EmpSize = ""
                Exp_Org_About = ""
                Exp_Org_Founded = ""

                # XPath for Organization Headline
                try:
                    Exp_Org_Headline_css_selector = 'h1.org-top-card-summary__title'
                    Exp_Org_Headline = driver.find_element(By.CSS_SELECTOR, Exp_Org_Headline_css_selector).text
                except NoSuchElementException:
                    Exp_Org_Headline = ""
                    print("Organization Headline not found")

                try:
                    Exp_Org_Industry = driver.find_element(By.CSS_SELECTOR,
                                                           'div.org-top-card-summary-info-list__info-item').text
                except NoSuchElementException:
                    Exp_Org_Industry = ""
                    print("Organization Industry not found")

                try:
                    # Use the CSS selector to find the elements
                    element_1 = driver.find_elements(By.CSS_SELECTOR,
                                                     'div.inline-block > .org-top-card-summary-info-list__info-item')

                    if len(element_1) >= 3:
                        Exp_Org_Location = element_1[0].text
                        Exp_Org_Followers = element_1[1].text.split()[0]
                        Exp_Org_EmpSize = element_1[2].text.split()[0]
                    else:
                        print("Not enough elements found for Location, Followers, or Employee Size")
                except NoSuchElementException:
                    Exp_Org_Location = ""
                    Exp_Org_Followers = ""
                    Exp_Org_EmpSize = ""
                    print("Organization details not found")

                try:
                    see_more = driver.find_element(By.CSS_SELECTOR, 'a.lt-line-clamp__more')
                    see_more.click()
                except:
                    pass

                try:
                    Exp_Org_About = driver.find_element(By.CSS_SELECTOR, 'div.org-about-module__description').text
                except:
                    Exp_Org_About = ""
                    print("Organization About section not found")

                url = a_tags_list[i] + "about/"
                try:
                    driver.get(url)
                    print("Entered about")
                    time.sleep(10)

                    # List of possible XPaths
                    xpaths = [
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[3]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[4]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[5]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[6]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[7]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[8]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[9]',
                        '/html/body/div[5]/div[3]/div/div[2]/div/div[2]/main/div[2]/div/div/div[1]/section/dl/dd[10]'
                    ]

                    Exp_Org_Founded = None

                    # Iterate over each XPath
                    for xpath in xpaths:
                        try:
                            # Try to find the element using the current XPath
                            Exp_Org_Founded = driver.find_element(By.XPATH, xpath).text
                            print(f"Found with XPath {xpath}: {Exp_Org_Founded}")
                            if Exp_Org_Founded:
                                try:
                                    # Attempt to convert the text to an integer
                                    int_value = int(Exp_Org_Founded)
                                    Exp_Org_Founded = int_value
                                    print(f"Exp_Org_Founded as integer: {Exp_Org_Founded}")
                                    break  # Exit the loop if found
                                except ValueError:
                                    # If conversion fails, store it as an empty string or custom message
                                    Exp_Org_Founded = "No Exp_Org_Founded"
                                    print("Conversion to integer failed, stored as 'No Exp_Org_Founded'")
                            else:
                                Exp_Org_Founded = "No Exp_Org_Founded"
                                print("Exp_Org_Founded not found, stored as 'No Exp_Org_Founded'")


                        except:
                            continue  # If not found, move to the next XPath



                except Exception as e:
                    Exp_Org_Founded = ""
                    print(f"An error occurred")
                # Printing the extracted values (for verification)
                print(f"Organization Headline: {Exp_Org_Headline}")
                print(f"Industry: {Exp_Org_Industry}")
                print(f"Location: {Exp_Org_Location}")
                print(f"Employee Size: {Exp_Org_EmpSize}")
                print(f"Followers: {Exp_Org_Followers}")
                print(f"About Section: {Exp_Org_About}")
                print(f"Founded Date: {Exp_Org_Founded}")

                company_data = {
                    "Exp_Org_Headline": Exp_Org_Headline,
                    "Exp_Org_Industry": Exp_Org_Industry,
                    "Exp_Org_Location": Exp_Org_Location,
                    "Exp_Org_EmpSize": Exp_Org_EmpSize,
                    "Exp_Org_Followers": Exp_Org_Followers,
                    "Exp_Org_About": Exp_Org_About,
                    "Exp_Org_Founded": Exp_Org_Founded,


                }
                # Add the dictionary to the list
                company_details.append(company_data)

            print(len(data_list))
            print(len(company_details))

            for i, data in enumerate(data_list):
                # Check if Exp_Org_HasLink is "1" or 1
                if data.get("Exp_Org_HasLink") in ["1", 1]:
                    # Add corresponding company details from company_details
                    data["company_details"] = company_details[i]
                else:
                    # Add empty string for company_details
                    data["company_details"] = " "
            #
            # json_data_company = json.dumps(company_details, ensure_ascii=False, indent=2)
                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

                # Write the JSON string to the Excel sheet
            sheet3.cell(row=row, column=current_col, value=json_data)
            # company_column = current_col + 1
            # sheet3.cell(row=row, column=company_column, value=json_data_company)
            print("JSON data written to sheet")

                # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")


        else:

            print("No experience found.")

    except NoSuchElementException as No_Such_Element_Found:
        print("No Experience data are extracted...")



    else:
        print("Experience data are extracted")


def education(driver, profile_url, sheet4, row):
    print(f"Navigating to profile education page")

    # Navigate to the profile URL
    driver.get(profile_url)

    try:

        try:
            # Locate "Show all licenses & certifications" link and navigate to it
            show_all_link = driver.find_element(By.CSS_SELECTOR,
                                                'a#navigation-index-see-all-education')
            show_all_url = show_all_link.get_attribute('href')
        except:
            show_all_url = profile_url + "details/education/"
            print(show_all_url)

        driver.get(show_all_url)
        # Set initial headers in the first row
        headers = ["LinkenIn_URL",
                   "Educ_Count",
                   "Education_Content"
                   ]

        data_list = []

        # Dynamically add headers for each educational entry (max 10 in this example)
        max_education_entries = 5

        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet4.cell(row=1, column=col, value=header)

        # Find all license & certification items
        education_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        Educ_Count = len(education_list)
        print(f"Number of education_list found: {Educ_Count}")
        sheet4.cell(row=row, column=1, value=profile_url)

        sheet4.cell(row=row, column=2, value=Educ_Count)
        #
        # links_list = []
        # a_tags = education_list.find_elements(By.XPATH,
        #                                          ".//li//a[contains(@href, '/school/') or contains(@href, '/company/') or contains(@href, 'search/results/all/') or contains(@href, 'keywords=') and contains(@href, 'institute')]")
        #
        # for a_tag in a_tags[::2]:  # Slice to get every other element
        #     link = a_tag.get_attribute("href")
        #     links_list.append(link)
        #
        # print(links_list)

        # Initialize the starting column for data entry
        current_col = 3

        # Write Educ_Count to the second row
        sheet4.cell(row=row, column=2, value=Educ_Count)

        for i, cert in enumerate(education_list[:5]):
            try:
                Educ_Org = cert.find_element(By.CSS_SELECTOR, '.display-flex .mr1.t-bold span').text.strip()
                print(Educ_Org)
            except Exception as e:
                Educ_Org = ""
                # print(f"Error retrieving Certif{i + 1}_Name: {e}")

            try:
                Educ_Degree_twice = cert.find_element(By.CSS_SELECTOR, '.t-14.t-normal').text
                Educ_Degree_list = Educ_Degree_twice.split('\n')
                Educ_Degree = Educ_Degree_list[0] if isinstance(Educ_Degree_list, list) else ""
                print(Educ_Degree)
            except Exception as e:
                Educ_Degree = ""
                # print(f"Error retrieving Certif{i + 1}_Org: {e}")

            try:
                Educ_Start_list = cert.find_element(By.CSS_SELECTOR, '.t-14.t-black--light').text.split('\n')
                Educ_Start = Educ_Start_list[0] if len(Educ_Start_list) > 0 else ""
                Educ_Start = Educ_Start.split(" - ")[0].strip()
                Educ_End = Educ_Start_list[0] if len(Educ_Start_list) > 0 else ""
                Educ_End = Educ_End.split(" - ")[1].strip()
                print(Educ_Start, Educ_End)
            except Exception as e:
                Educ_Start = ""
                Educ_End = ""
                # print(f"Error retrieving Certif{i + 1}_Issue_Date: {e}")

            try:
                Educ_Org_Link = cert.find_element(By.CSS_SELECTOR,
                                                  "li a.optional-action-target-wrapper").get_attribute("href")

                # if Educ_Org_Link and any(
                #         keyword in Educ_Org_Link for keyword in ["company", "school", "institution"]):
                #     Exp_Org_HasLink = "1"
                # else:
                #     Exp_Org_HasLink = "0"
                Educ_Org_Has_Link = "1" if Educ_Org_Link else "0"

                print(Educ_Org_Link, Educ_Org_Has_Link)
            except Exception as e:
                Educ_Org_Link = ""
                Educ_Org_Has_Link = "0"
                # print(f"Error retrieving Certif{i + 1}_Org_Link: {e}")


            # current_col += 6  # Move to the next set of columns for the next entry
            # Append extracted data to the list
            Education_data = {
                "Educ_Count": i + 1,
                "Educ_Org": Educ_Org,
                "Educ_Degree": Educ_Degree,
                "Educ_Start": Educ_Start,
                "Educ_End": Educ_End,
                "Educ_Org_Has_Link": Educ_Org_Has_Link,
                "Educ_Org_Link": Educ_Org_Link
            }

            data_list.append(Education_data)
            print(f"Post {i + 1} data appended successfully.")

            # Convert the data list to a JSON string
        json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

        # Write the JSON string to the Excel sheet
        sheet4.cell(row=row, column=current_col, value=json_data)
        print("JSON data written to sheet")

        # Save the workbook
        workbook.save("profile_basic_data_7.xlsx")
        print("File saved successfully")
    except:
        print("No Education data are extracted")


    else:
        print("Education data are extracted")


def licenses_certifications(driver, profile_url, sheet5, row):
    print(profile_url)
    driver.get(profile_url)
    print("License & certifications data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Certif_Count","license_certifications_content"]

        # Locate "Show all licenses & certifications" link and navigate to it
        # show_all_link = driver.find_element(By.CSS_SELECTOR, 'a#navigation-index-see-all-licenses-and-certifications')
        # show_all_url = show_all_link.get_attribute('href')
        # print(show_all_url)
        # driver.get(show_all_url)
        try:
            # Locate "Show all licenses & certifications" link and navigate to it
            show_all_link = driver.find_element(By.CSS_SELECTOR,
                                                'a#navigation-index-see-all-licenses-and-certifications')
            show_all_url = show_all_link.get_attribute('href')
        except:
            show_all_url = profile_url + "details/certifications/"
            print(show_all_url)

        driver.get(show_all_url)

        # Find all license & certification items
        license_certification_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        Certif_Count = len(license_certification_list)
        print(f"Number of licenses & certifications found: {Certif_Count}")
        entries = 5


        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet5.cell(row=1, column=col, value=header)

        sheet5.cell(row=row, column=1, value=profile_url)

        sheet5.cell(row=row, column=2, value=Certif_Count)

        # If certificates are found, extract the data
        if Certif_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for Certif_Count)
            data_list = []

            for i, cert in enumerate(license_certification_list[:5]):
                try:
                    Certif_Name = cert.find_element(By.CSS_SELECTOR, '.display-flex .mr1.t-bold span').text.strip()
                except Exception as e:
                    Certif_Name = ""
                    # print(f"Error retrieving Certif{i + 1}_Name: {e}")

                try:
                    Certif_Org_twice = cert.find_element(By.CSS_SELECTOR, '.t-14.t-normal').text
                    Certif_Org_list = Certif_Org_twice.split('\n')
                    Certif_Org = Certif_Org_list[0] if isinstance(Certif_Org_list, list) else ""
                except Exception as e:
                    Certif_Org = ""

                try:
                    Certif_Issue_Date_list = cert.find_element(By.CSS_SELECTOR, '.t-14.t-black--light').text.split('\n')
                    Certif_Issue_Date = Certif_Issue_Date_list[0] if len(Certif_Issue_Date_list) > 0 else ""
                except Exception as e:
                    Certif_Issue_Date = ""

                try:
                    Certif_Skills_css_selector = "li.pvs-list__item--with-top-padding"
                    Certif_Skills_element = cert.find_element(By.CSS_SELECTOR, Certif_Skills_css_selector).text
                    Certif_Skills = Certif_Skills_element
                    print(Certif_Skills_element)
                    # Certif_Skills = Certif_Skills_element.text.strip() if Certif_Skills_element else "Skill not mentioned"
                except Exception as e:
                    Certif_Skills = "Skill not mentioned"

                try:
                    Certif_Org_Link = cert.find_element(By.CSS_SELECTOR,
                                                        "li a.optional-action-target-wrapper").get_attribute("href")

                    Certif_Org_Has_Link = "1" if Certif_Org_Link else "0"
                except Exception as e:
                    Certif_Org_Link = ""
                    Certif_Org_Has_Link = "0"

                Certification_data = {
                    "Certif_Count": i + 1,
                    "Certif_Name": Certif_Name,
                    "Certif_Org": Certif_Org,
                    "Certif_Issue_Date": Certif_Issue_Date,
                    "Certif_Skills": Certif_Skills,
                    "Certif_Org_Has_Link": Certif_Org_Has_Link,
                    "Certif_Org_Link": Certif_Org_Link
                }

                data_list.append(Certification_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet5.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No licenses & certifications found.")

    except NoSuchElementException as No_Such_Element_Found:
        print("No License & Certifications data are extracted")




    else:
        print("License & Certifications data are extracted")


def projects(driver, profile_url, sheet6, row):
    print(profile_url)
    driver.get(profile_url)
    print("Projects data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Project_Count", "projects_content"]

        # Locate "Show all projects" link and navigate to it
        try:
            show_all_link = driver.find_element(By.CSS_SELECTOR, 'a#navigation-index-see-all-projects')
            show_all_url = show_all_link.get_attribute('href')
        except:
            show_all_url = profile_url + "details/projects/"
            print(show_all_url)
        driver.get(show_all_url)

        # Find all license & certification items
        projects_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        Project_Count = len(projects_list)
        print(f"Number of projects found: {Project_Count}")
        entries = 6


        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet6.cell(row=1, column=col, value=header)

        sheet6.cell(row=row, column=1, value=profile_url)

        sheet6.cell(row=row, column=2, value=Project_Count)

        if show_all_url:
            print("projects show_all_url found.")
        else:
            print("projects show_all_url not found.")

        # If certificates are found, extract the data
        if Project_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for Certif_Count)
            data_list = []

            for i, cert in enumerate(projects_list[:5]):
                entry_text = cert.text.split("\n")  # Split the text by newline
                print(entry_text)
                if "Nothing to see for now" in entry_text:
                    sheet6.cell(row=row, column=2, value=0)
                    Project_Name = ""
                    Project_Desc = ""
                    Project_Start = ""
                    Project_End = ""
                    Project_Org = ""
                    sheet6.cell(row=row, column=current_col, value=0)
                    break

                else:
                    try:
                        Project_Name = entry_text[0] if len(entry_text) > 0 else ""
                    except Exception:
                        Project_Name = ""

                    try:
                        Project_Desc = entry_text[7] if len(entry_text) > 6 else ""
                    except Exception:
                        Project_Desc = ""

                    try:
                        Project_Start = entry_text[2] if len(entry_text) > 2 else ""
                        Project_Start = Project_Start.split(" - ")[0].strip()
                    except Exception:
                        Project_Start = ""

                    try:
                        Project_End = entry_text[2] if len(entry_text) > 2 else ""
                        Project_End = Project_End.split(" - ")[1].strip()
                    except Exception:
                        Project_End = ""

                    try:
                        Project_Org = entry_text[4] if len(entry_text) > 4 else ""
                    except Exception:
                        Project_Org = ""

                project_data = {
                    "project_Count": i + 1,
                    "Project_Name": Project_Name,
                    "Project_Desc": Project_Desc,
                    "Project_Start": Project_Start,
                    "Project_End": Project_End,
                    "Project_Org": Project_Org,

                }

                data_list.append(project_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet6.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No projects found.")

    except NoSuchElementException as No_Such_Element_Found:
        print("Entered except 1 part...")
        print("Projects data are extracted")


    else:

        print("Projects data are extracted")


def volunteering(driver, profile_url, sheet7, row):
    print(profile_url)
    # driver.get(profile_url)
    print("volunteering data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "volunteering_Count", "Volunteering_content"]

        # Locate "Show all volunteerings" link and navigate to it
        try:
            show_all_link = driver.find_element(By.CSS_SELECTOR, 'a#navigation-index-see-all-volunteer-experiences')
            show_all_url = show_all_link.get_attribute('href')
        except:

            show_all_url = profile_url + "details/volunteering-experiences/"
            print(show_all_url)
        driver.get(show_all_url)

        # Find all license & certification items
        volunteering_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        volunteering_Count = len(volunteering_list)
        print(f"Number of volunteering found: {volunteering_Count}")
        entries = 5


        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet7.cell(row=1, column=col, value=header)

        sheet7.cell(row=row, column=1, value=profile_url)

        sheet7.cell(row=row, column=2, value=volunteering_Count)

        if show_all_url:
            print("volunteerings show_all_url found.")
        else:
            print("volunteerings show_all_url not found.")

        # If certificates are found, extract the data
        if volunteering_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for volunteering_Count)
            data_list = []
            for i, cert in enumerate(volunteering_list[:5]):
                print("Entered Volunteering for loop")
                entry_text = cert.text.split("\n")  # Split the text by newline
                print(entry_text)

                if "Nothing to see for now" in entry_text:
                    sheet7.cell(row=row, column=2, value=0)
                    Volunteer_Org_Name = ""
                    Volunteer_Title = ""
                    Volunteer_Cause = ""
                    Volunteer_Start = ""
                    Volunteer_End = ""
                    Volunteer_Desc = ""
                    Volunteer_Org_Link = ""
                    Volunteer_Org_HasLink = "0"
                    sheet7.cell(row=row, column=current_col, value=0)
                    break


                else:
                    try:
                        Volunteer_Org_Name = entry_text[2] if len(entry_text) > 2 else ""
                    except:
                        Volunteer_Org_Name = ""

                    try:
                        Volunteer_Title = entry_text[0] if len(entry_text) > 0 else ""
                    except:
                        Volunteer_Title = ""

                    try:
                        Volunteer_Cause = entry_text[6] if len(entry_text) > 6 else ""
                    except:
                        Volunteer_Cause = ""

                    try:
                        Volunteer_Start = entry_text[4] if len(entry_text) > 4 else ""
                        Volunteer_Start = Volunteer_Start.split(" - ")[0].strip()
                    except:
                        Volunteer_Start = ""

                    try:
                        Volunteer_End = entry_text[4] if len(entry_text) > 4 else ""
                        Volunteer_End_ = Volunteer_End.split(" - ")[1].strip()
                        print(Volunteer_End_)
                        Volunteer_End = Volunteer_End_.split("·")[0].strip()
                        print(Volunteer_End)

                    except:
                        Volunteer_End = ""

                    try:
                        Volunteer_Desc = entry_text[8] if len(entry_text) > 8 else ""
                    except:
                        Volunteer_Desc = ""

                    try:
                        Volunteer_Org_Link = cert.find_element(By.CSS_SELECTOR,
                                                               "li a.optional-action-target-wrapper").get_attribute(
                            "href")
                        Volunteer_Org_HasLink = "1" if Volunteer_Org_Link else "0"
                    except:
                        Volunteer_Org_Link = ""
                        Volunteer_Org_HasLink = "0"

                Volunteer_data = {
                    "Volunteer_Count": i + 1,
                    "Volunteer_Org_Name": Volunteer_Org_Name,
                    "Volunteer_Title": Volunteer_Title,
                    "Volunteer_Cause": Volunteer_Cause,
                    "Volunteer_Start": Volunteer_Start,
                    "Volunteer_End": Volunteer_End,
                    "Volunteer_Desc": Volunteer_Desc,
                    "Volunteer_Org_Link": Volunteer_Org_Link,
                    "Volunteer_Org_HasLink": Volunteer_Org_HasLink,

                }


                data_list.append(Volunteer_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet7.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No volunteerings found.")

    except NoSuchElementException as No_Such_Element_Found:
        print("Entered except 1 part...")
        print("Volunteering data are extracted")



    else:
        print("Volunteering data are extracted")


def honors(driver, profile_url, sheet8, row):
    driver.get(profile_url)
    # Honors_Count = ""
    # Honors1_Title = ""
    # Honor1_Org = ""
    # Honors1_Issuer = ""
    # Honors1_Issue_Date = ""
    # Honors1_Desc = ""

    print(profile_url)

    print("Honors data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Honors_Count", "Honors_Count"]

        # Locate "Show all Honors" link and navigate to it
        try:
            show_all_link = driver.find_element(By.CSS_SELECTOR, 'a#navigation-index-see-all-honorsandawards')
            show_all_url = show_all_link.get_attribute('href')
        except:

            show_all_url = profile_url + "details/honors/"
            print(show_all_url)
        driver.get(show_all_url)

        # Find all license & certification items
        honor_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        Honors_Count = len(honor_list)
        print(f"Number of honor found: {Honors_Count}")
        entries = 6


        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet8.cell(row=1, column=col, value=header)
        sheet8.cell(row=row, column=1, value=profile_url)

        sheet8.cell(row=row, column=2, value=Honors_Count)

        if show_all_url:
            print("honors show_all_url found.")
        else:
            print("honors show_all_url not found.")
        current_col = 3
        # If certificates are found, extract the data
        if Honors_Count > 0:
              # Start from column B for the first certificate (Column A is for honor_Count)
            data_list = []
            for i, cert in enumerate(honor_list[:5]):
                entry_text = cert.text.split("\n")  # Split the text by newline
                print(entry_text)
                if "Nothing to see for now" in entry_text:
                    sheet8.cell(row=row, column=2, value=0)

                    Honors_Title = ""
                    Honors_Org = ""
                    Honors_Issuer = ""
                    Honors_Issue_Date = ""
                    Honors_Desc = ""
                    sheet8.cell(row=row, column=current_col, value=0)
                    break
                else:

                    try:
                        Honors_Title = entry_text[0] if len(entry_text) > 0 else ""
                    except Exception:
                        Honors_Title = ""

                    try:
                        Honors_Org = entry_text[4] if len(entry_text) > 4 else ""
                    except Exception:
                        Honors_Org = ""

                    try:
                        Honors_Issuer = entry_text[2] if len(entry_text) > 2 else ""
                        Honors_Issuer = Honors_Issuer.split("·")[0].strip()
                    except Exception:
                        Honors_Issuer = ""

                    try:
                        Honors_Issue_Date = entry_text[2] if len(entry_text) > 2 else ""
                        Honors_Issue_Date = Honors_Issue_Date.split("·")[1].strip()
                    except Exception:
                        Honors_Issue_Date = ""

                    try:
                        Honors_Desc = entry_text[6] if len(entry_text) > 6 else ""
                    except Exception:
                        Honors_Desc = ""

                Honors_data = {

                    "Honors_Count": i + 1,
                    "Honors_Title": Honors_Title,
                    "Honors_Org": Honors_Org,
                    "Honors_Issuer": Honors_Issuer,
                    "Honors_Issue_Date": Honors_Issue_Date,
                    "Honors_Desc": Honors_Desc,


                }

                data_list.append(Honors_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet8.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No honors found.")

    except NoSuchElementException as No_Such_Element_Found:
        print("Entered except 1 part...")
        print("No Honors data are extracted")




    else:

        print("Honors data are extracted")


def skills(driver, profile_url, sheet9, row):
    driver.get(profile_url)
    # Skills_Count = ""
    # Skill1_Name = ""
    # Skill1_Endorse_Num = ""

    print(profile_url)

    print("Skills data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Skills_Count", "skills_content"]
        #
        # # Locate "Show all Skills" link and navigate to it
        # # try:
        # show_all_link_path = 'a#navigation-index-Show-all-18-skills'
        # show_all_link = driver.find_element(By.CSS_SELECTOR, show_all_link_path)
        # # show_all_url = show_all_link.get_attribute('href')
        # # except:
        # #
        show_all_url = profile_url + "details/skills/"
        print(show_all_url)
        driver.get(show_all_url)

        # Find all license & certification items
        Skill_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        Skills_Count = len(Skill_list)
        print(f"Number of Skill found: {Skills_Count}")
        entries = 5


        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet9.cell(row=1, column=col, value=header)

        sheet9.cell(row=row, column=1, value=profile_url)

        sheet9.cell(row=row, column=2, value=Skills_Count)

        if show_all_url:
            print("Skills show_all_url found.")
        else:
            print("Skills show_all_url not found.")

        # If certificates are found, extract the data
        if Skills_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for Skill_Count)
            data_list = []
            for i, entry in enumerate(Skill_list[:5]):

                if entry.text == '':
                    sheet9.cell(row=row, col=current_col, value=0)
                    break
                print(i)
                # print(entry.text)
                entry_text = entry.text.split("\n")  # Split the text by newline
                print(entry_text)

                try:
                    Skill_Name = entry_text[0] if len(entry_text) > 0 else ""
                except Exception:
                    Skill_Name = ""

                Skill_Endorse_Num = None

                try:
                    # Check if "endorsements" exists as a separate word in any element
                    for item in entry_text:
                        words = item.split()  # Split the item by spaces
                        if "endorsements" in words:
                            Skill_Endorse_Num = item.split()[0].strip()
                            break
                except Exception:
                    Skill_Endorse_Num = None

                print(Skill_Name)
                print(Skill_Endorse_Num)
                #
                skills_data = {

                    "Skill_Count": i + 1,
                    "Skill_Name": Skill_Name,
                    "Skill_Endorse_Num": Skill_Endorse_Num,


                }

                data_list.append(skills_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet9.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No Skills found.")

    except:
        print("Entered except 1 part...")
        print("No Skills data are extracted")


    else:
        print("Skills data are extracted")


def recommendations_received(driver, profile_url, sheet10, row):
    driver.get(profile_url)
    Rec_Received_Count = ""
    Rec_Received1_Name = ""
    Rec_Received1_Title = ""
    Rec_Received1_Date = ""
    Rec_Received1_Relation = ""
    Rec_Received1_Desc = ""

    print(profile_url)

    print("recommendations data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Rec_Received_Count", "recommendations_received_content"]

        # Locate "Show all recommendations_receiveds" link and navigate to it
        try:
            show_all_link = driver.find_element(By.CSS_SELECTOR, 'a#navigation-index-see-all-recommendations')
            show_all_url = show_all_link.get_attribute('href')
        except:

            show_all_url = profile_url + "details/recommendations/"
            print(show_all_url)
        driver.get(show_all_url)
        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet10.cell(row=1, column=col, value=header)
        sheet10.cell(row=row, column=1, value=profile_url)


        # Locate the active tab panel using its role and aria-selected attribute
        active_tab_panel = driver.find_element(By.CSS_SELECTOR, 'div[role="tabpanel"].artdeco-tabpanel.active')
        # Locate all "license & certification" items within the active tab panel
        recommendations_received_list = active_tab_panel.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')

        # # Find all license & certification items
        # recommendations_received_list = driver.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        recommendations_receiveds_Count = len(recommendations_received_list)
        print(f"Number of recommendations_received found: {recommendations_receiveds_Count}")
        entries = 5
        sheet10.cell(row=row, column=2, value=recommendations_receiveds_Count)

        if show_all_url:
            print("recommendations_receiveds show_all_url found.")
        else:
            print("recommendations_receiveds show_all_url not found.")

        # If certificates are found, extract the data
        if recommendations_receiveds_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for recommendations_received_Count)
            data_list = []
            for i, cert in enumerate(recommendations_received_list[:5]):
                # print(cert.text)
                entry_text = cert.text
                try:
                    # Split the entry text into lines and remove duplicates
                    lines = list(dict.fromkeys(entry_text.strip().split('\n')))

                    # Remove any lines that contain irrelevant keywords like 'profile picture'
                    filtered_lines = [line for line in lines if "profile picture" not in line.lower()]

                    try:
                        Rec_Received_Name = filtered_lines[0].strip()
                    except Exception:
                        Rec_Received_Name = ""

                    try:
                        Rec_Received_Title = filtered_lines[1].strip()
                    except Exception:
                        Rec_Received_Title = ""

                    try:
                        Rec_Received_Date = filtered_lines[2].split(",")[0].strip()  # Extract the date
                    except Exception:
                        Rec_Received_Date = ""

                    try:
                        Rec_Received_year = filtered_lines[2].split(",")[1].strip()
                    except Exception:
                        Rec_Received_year = ""

                    # Combine the date and year
                    Rec_Received_Date = Rec_Received_Date + " " + Rec_Received_year

                    try:
                        Rec_Received_Relation = filtered_lines[2].split(",")[2].strip()  # Extract the relationship
                        if not Rec_Received_Relation:
                            Rec_Received_Relation = "No Relation mentioned"
                    except Exception:
                        Rec_Received_Relation = "No Relation mentioned"

                    try:
                        Rec_Received_Desc = max(filtered_lines,
                                                key=len)  # Combine all remaining lines as the description
                    except Exception:
                        Rec_Received_Desc = ""

                except:
                    Rec_Received_Name = ""
                    Rec_Received_Title = ""
                    Rec_Received_Date = ""
                    Rec_Received_Relation = ""
                    Rec_Received_Desc = ""
                    sheet10.cell(row=row, column=current_col, value=0)

                Rec_Received_data = {

                    "Rec_Received__Count": i + 1,
                    "Rec_Received_Name": Rec_Received_Name,
                    "Rec_Received_Title": Rec_Received_Title,
                    "Rec_Received_Date": Rec_Received_Date,
                    "Rec_Received_Relation": Rec_Received_Relation,
                    "Rec_Received_Desc": Rec_Received_Desc,

                }

                data_list.append(Rec_Received_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet10.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:

            print("No recommendations_receiveds found.")

    except NoSuchElementException as No_Such_Element_Found:
        print("Entered except 1 part...")
        print("No Recommendations Received data are extracted")



    else:

        print("Recommendations Received data are extracted")


def recommendations_given(driver, profile_url, sheet11, row):
    driver.get(profile_url)
    # Rec_Given_Count = ""
    # Rec_Given1_Name = ""
    # Rec_Given1_Title = ""
    # Rec_Given1_Date = ""
    # Rec_Given1_Relation = ""
    # Rec_Given1_Desc = ""

    print(profile_url)

    print("recommendations data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Rec_Given_Count", "Recommendations_given_content"]

        # Locate "Show all recommendations_Givens" link and navigate to it
        try:
            show_all_link = driver.find_element(By.CSS_SELECTOR, 'a#navigation-index-see-all-recommendations')
            show_all_url = show_all_link.get_attribute('href')
        except:
            #
            show_all_url = profile_url + "details/recommendations/"
            print(show_all_url)
        driver.get(show_all_url)
        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet11.cell(row=1, column=col, value=header)
        sheet11.cell(row=row, column=1, value=profile_url)


        # Find the "Given" button using XPath and click it
        given_button = driver.find_element(By.XPATH,
                                           "//button[contains(@id, 'ember') and contains(@aria-controls, 'ember')][span[text()='Given']]")
        given_button.click()

        # Locate the active tab panel using its role and aria-selected attribute
        active_tab_panel = driver.find_element(By.CSS_SELECTOR, 'div[role="tabpanel"].artdeco-tabpanel.active')

        # Find all license & certification items
        recommendations_Given_list = active_tab_panel.find_elements(By.CLASS_NAME, 'pvs-list__paged-list-item')
        recommendations_Givens_Count = len(recommendations_Given_list)
        print(f"Number of recommendations_Given found: {recommendations_Givens_Count}")
        entries = 5
        sheet11.cell(row=row, column=2, value=recommendations_Givens_Count)

        if show_all_url:
            print("recommendations_Givens show_all_url found.")
        else:
            print("recommendations_Givens show_all_url not found.")

        # If certificates are found, extract the data
        if recommendations_Givens_Count > 0:
            current_col = 3  # Start from column B for the first certificate (Column A is for recommendations_Given_Count)
            data_list = []
            for i, cert in enumerate(recommendations_Given_list[:5]):
                print(cert.text)
                entry_text = cert.text

                if "You haven't written any recommendations yet" in entry_text:
                    sheet11.cell(row=row, col=current_col, value = 0)
                    break


                try:
                    # Split the entry text into lines and remove duplicates
                    lines = list(dict.fromkeys(entry_text.strip().split('\n')))
                    #
                    # # Remove any lines that contain irrelevant keywords like 'profile picture'
                    # Remove any lines that contain irrelevant keywords like 'profile picture'
                    filtered_lines = [line for line in lines if "profile picture" not in line.lower()]

                    try:
                        Rec_Given_Name = filtered_lines[0].strip()
                    except Exception:
                        Rec_Given_Name = ""

                    try:
                        Rec_Given_Title = filtered_lines[1].strip()
                    except Exception:
                        Rec_Given_Title = ""

                    try:
                        Rec_Given_Date = filtered_lines[2].split(",")[0].strip()  # Extract the date
                    except Exception:
                        Rec_Given_Date = ""

                    try:
                        Rec_Given_year = filtered_lines[2].split(",")[1].strip()
                    except Exception:
                        Rec_Given_year = ""

                    # Combine the date and year
                    Rec_Given_Date = Rec_Given_Date + " " + Rec_Given_year

                    try:
                        Rec_Given_Relation = filtered_lines[2].split(",")[2].strip()  # Extract the relationship
                        if not Rec_Given_Relation:
                            Rec_Given_Relation = "No Relation mentioned"
                    except Exception:
                        Rec_Given_Relation = "No Relation mentioned"

                    try:
                        Rec_Given_Desc = max(filtered_lines, key=len)  # Combine all remaining lines as the description
                    except Exception:
                        Rec_Given_Desc = ""
                except:

                    Rec_Given_Name = ""
                    Rec_Given_Title = ""
                    Rec_Given_Date = ""
                    Rec_Given_Relation = ""
                    Rec_Given_Desc = ""
                    sheet11.cell(row=row, column=current_col, value=0)

                Rec_Given_data = {

                    "Rec_Given__Count": i + 1,
                    "Rec_Given_Name": Rec_Given_Name,
                    "Rec_Given_Title": Rec_Given_Title,
                    "Rec_Given_Date": Rec_Given_Date,
                    "Rec_Given_Relation": Rec_Given_Relation,
                    "Rec_Given_Desc": Rec_Given_Desc,

                }

                data_list.append(Rec_Given_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet11.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No recommendations_Givens found.")

    except NoSuchElementException as No_Such_Element_Found:

        print("Entered except 1 part...")
        print("No Recommendations Given data are extracted")



    else:
        print("Recommendations Given data are extracted")


def feature_post(driver, profile_url, sheet12, row):
    driver.get(profile_url)
    Featured_Count = ""
    Featured_Type = ""
    Featured_Title = ""
    Featured_Desc = ""
    print(profile_url)

    print("Feature post data are getting extracted...")

    try:
        print("Entered try block...")

        # Set initial headers in the first row
        headers = ["LinkenIn_URL", "Featured_Count", "Featured_Content"]

        show_all_url = profile_url + "details/featured/"
        print(show_all_url)
        driver.get(show_all_url)

        # Find all license & certification items
        featured_list = driver.find_elements(By.CSS_SELECTOR,
                                             '.pvs-list__paged-list-item.pvs-list__item--with-gap.pvs-list__item--one-column')
        featureds_Count = len(featured_list)
        Featured_Count = len(featured_list)
        print(f"Number of featured found: {featureds_Count}")
        entries = 5


        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet12.cell(row=1, column=col, value=header)
        sheet12.cell(row=row, column=1, value=profile_url)

        sheet12.cell(row=row, column=2, value=Featured_Count)

        if show_all_url:
            print("featureds show_all_url found.")
        else:
            print("featureds show_all_url not found.")
        current_col = 3
        # If certificates are found, extract the data
        if featureds_Count > 0:
              # Start from column B for the first certificate (Column A is for featured_Count)
            data_list = []
            for i, cert in enumerate(featured_list[:5]):
                entry_text = cert.text
                if "Nothing to see for now" in entry_text:
                    sheet12.cell(row=row, column=2, value=0)
                    Featured_Type = ""
                    Featured_Title = ""
                    Featured_Desc = ""
                    sheet12.cell(row=row, column=current_col, value=0)
                    break
                else:
                    try:
                        featured_content_list = list(dict.fromkeys(entry_text.strip().split('\n')))
                        print(featured_content_list)

                        try:
                            Featured_Type = featured_content_list[0]
                        except Exception:
                            Featured_Type = ""

                        try:
                            Featured_Title = featured_content_list[1]
                        except Exception:
                            Featured_Title = ""

                        try:
                            Featured_Desc = max(featured_content_list, key=len)
                        except Exception:
                            Featured_Desc = ""

                    except:
                        Featured_Type = ""
                        Featured_Title = ""
                        Featured_Desc = ""

                Featured_data = {

                    "Featured__Count": i + 1,
                    "Featured_Type": Featured_Type,
                    "Featured_Title": Featured_Title,
                    "Featured_Desc": Featured_Desc,

                }

                data_list.append(Featured_data)
                print(f"Post {i + 1} data appended successfully.")

                # Convert the data list to a JSON string
            json_data = json.dumps(data_list, ensure_ascii=False, indent=2)

            # Write the JSON string to the Excel sheet
            sheet12.cell(row=row, column=current_col, value=json_data)
            print("JSON data written to sheet")

            # Save the workbook
            workbook.save("profile_basic_data_7.xlsx")
            print("File saved successfully")
        else:
            print("No featured found.")
            sheet12.cell(row=row, column=current_col, value=0)

    except NoSuchElementException as No_Such_Element_Found:

        print("Couldn't navigate to the page")
    else:
        print("Recommendations Given data are extracted")


def main():
    # Start timing
    start_time = time.time()
    # Execution
    print("Starting LinkedIn login...")
    linkedIn_login(driver)
    time.sleep(30)
    row = 2

    for link in links:
        profile_url = link
        print(profile_url)

        row = row
        #
        print("Extracting profile basic data...")
        profile_basic_data(driver, profile_url, sheet1, row)
        time.sleep(5)
        #
        print("Extracting activity data...")
        activity(driver, profile_url, sheet2, row)
        time.sleep(5)
        # #
        # 
        # 
        # print("Extracting experience data...")
        # experience(driver, profile_url, sheet3, row)
        # #
        # time.sleep(5)
        # # #
        # print("Extracting education data...")
        # education(driver, profile_url, sheet4, row)
        # time.sleep(5)
        # 
        # 
        # print("Extracting License&Certifications details...")
        # licenses_certifications(driver, profile_url, sheet5, row)
        # time.sleep(5)
        # 
        # # #
        # print("Extracting Projects details...")
        # projects(driver, profile_url, sheet6, row)
        # time.sleep(5)
        # #
        # #
        # print("Extracting Volunteering details...")
        # volunteering(driver, profile_url, sheet7, row)
        # time.sleep(5)
        # #
        # #
        # print("Extracting Honors details...")
        # honors(driver, profile_url, sheet8, row)
        # time.sleep(5)
        # #
        # #
        # print("Extracting Skills details...")
        # skills(driver, profile_url, sheet9, row)
        # time.sleep(5)
        # #
        # #
        # print("Extracting Recommendations Received details...")
        # recommendations_received(driver, profile_url, sheet10, row)
        # time.sleep(5)
        # #
        # # #
        # print("Extracting Recommendations Given details...")
        # recommendations_given(driver, profile_url, sheet11, row)
        # time.sleep(5)
        # 
        # 
        # print("Extracting featured post data...")
        # feature_post(driver, profile_url, sheet12, row)
        # time.sleep(5)
        #
        row = row + 1

    print("Closing the browser...")
    driver.quit()

    # End timing
    end_time = time.time()

    # Calculate and print the time taken
    elapsed_time = end_time - start_time
    print(f"Time taken to run the program: {elapsed_time:.2f} seconds")


if __name__ == "__main__":
    main()
